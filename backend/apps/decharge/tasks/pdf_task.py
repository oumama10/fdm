"""
Celery task – generate a signed PDF décharge from the official Excel template.

Queue  : pdf
Retries: 2  (with exponential backoff)

Template cell references
------------------------
All cell coordinates and the first data row are defined as module-level
constants so they can be updated in one place when the real template is
supplied without touching any other logic.

Template expected layout
^^^^^^^^^^^^^^^^^^^^^^^^
Row 1-9  : header block (numero, date, service …)
Row 10+  : article table (one article per row)

Adjust the constants below to match the production template.
"""

from __future__ import annotations

import logging
import os
import shutil
import subprocess
import tempfile
import uuid
from pathlib import Path

from celery.exceptions import MaxRetriesExceededError
from django.conf import settings

from config.celery import app

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Template constants — adjust to match the real .xlsx layout
# ---------------------------------------------------------------------------

# Header cells (openpyxl A1-style coordinates)
CELL_NUMERO_DECHARGE: str = "C3"   # N° décharge
CELL_DATE: str = "C4"              # Date de génération
CELL_SERVICE: str = "C5"           # Service destinataire

# First data row (1-based) for the articles table
ROW_START: int = 10

# Column indices (1-based) for the articles table
COL_INDEX: int = 1        # A — row number
COL_DESIGNATION: int = 2  # B — article designation
COL_QUANTITE: int = 3     # C — quantity
COL_INVENTAIRE: int = 4   # D — inventory number (bien inventaire only)

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _notify_gestionnaires(msg_titre: str, msg_body: str, ref_obj=None) -> None:
    """
    Send a *web* Notification to every active gestionnaire_magasin.
    Silently swallows exceptions so it never causes the main task to fail.
    """
    try:
        from apps.alerts.models import Notification          # noqa: PLC0415
        from apps.users.models import Utilisateur            # noqa: PLC0415
        from django.contrib.contenttypes.models import ContentType  # noqa: PLC0415

        gestionnaires = list(
            Utilisateur.objects.filter(
                id_role__nom_role="gestionnaire_magasin",
                actif=True,
            ).only("id_utilisateur")
        )
        if not gestionnaires:
            return

        ct = None
        obj_id = None
        if ref_obj is not None:
            ct = ContentType.objects.get_for_model(ref_obj)
            obj_id = ref_obj.pk

        Notification.objects.bulk_create(
            [
                Notification(
                    id_destinataire=g,
                    type_notification="validation_requise",
                    titre=msg_titre,
                    message=msg_body,
                    canal="web",
                    content_type=ct,
                    object_id=obj_id,
                )
                for g in gestionnaires
            ]
        )
    except Exception:
        logger.exception(
            "generate_decharge_pdf: failed to notify gestionnaires ('%s')",
            msg_titre,
        )


# ---------------------------------------------------------------------------
# Celery task
# ---------------------------------------------------------------------------


@app.task(queue="pdf", bind=True, max_retries=2)
def generate_decharge_pdf(self, decharge_id: int) -> None:
    """
    Generate the signed PDF décharge for *Decharge(pk=decharge_id)*.

    Steps
    -----
    1.  Fetch Decharge with all related data (select_related + prefetch_related).
    2.  Open the Excel template from ``settings.DECHARGE_TEMPLATE_PATH``
        (openpyxl, keep_vba=False, data_only=False).
    3.  Fill header cells (numero, date, service).
    4.  Write one row per LigneDecharge starting at ROW_START.
    5.  Save the filled workbook to a temp .xlsx file under /tmp/.
    6.  Run LibreOffice headless → PDF conversion (timeout 60 s).
    7.  Move the resulting PDF to MEDIA_ROOT/decharges/pdf/.
    8.  Update Decharge.fichier_pdf with the relative media path.
    9.  Create SignatureDecharge(statut='en_attente') if not already present.
    10. Send 'decharge_prete' Notification to the chef_service.
    11. Clean up the temp .xlsx file.
    12. On any exception: retry with exponential backoff (60 s → 120 s).
        After max_retries: send error Notification to all gestionnaires.
    """
    # All model imports are deferred so this module can be loaded before the
    # Django app registry is ready (e.g. during Celery worker boot).
    from apps.decharge.models import Decharge, SignatureDecharge  # noqa: PLC0415
    from apps.alerts.models import Notification                   # noqa: PLC0415
    from django.contrib.contenttypes.models import ContentType    # noqa: PLC0415

    import openpyxl                                               # noqa: PLC0415

    xlsx_tmp_path: Path | None = None  # tracked for cleanup regardless of outcome

    try:
        # ── 1. Fetch Decharge with all required related objects ───────────────
        try:
            decharge = (
                Decharge.objects.select_related(
                    "id_demande__id_service",
                    "id_demande__id_chef_demandeur",
                )
                .prefetch_related(
                    "lignes__id_ressource",
                    "lignes__id_instance_ressource",
                )
                .get(pk=decharge_id)
            )
        except Decharge.DoesNotExist:
            logger.error(
                "generate_decharge_pdf: Decharge pk=%s not found", decharge_id
            )
            return  # permanent failure — no point retrying a missing object

        demande = decharge.id_demande
        service_nom = (
            demande.id_service.nom_service if demande.id_service else "—"
        )
        chef = demande.id_chef_demandeur

        # ── 2. Open the Excel template ────────────────────────────────────────
        template_path = Path(settings.DECHARGE_TEMPLATE_PATH)
        if not template_path.is_file():
            raise FileNotFoundError(
                f"DECHARGE_TEMPLATE_PATH does not exist: {template_path}"
            )

        wb = openpyxl.load_workbook(
            template_path,
            keep_vba=False,
            data_only=False,
        )

        # ── 3. Access active sheet and fill header ────────────────────────────
        ws = wb.active

        ws[CELL_NUMERO_DECHARGE] = decharge.numero_decharge
        ws[CELL_DATE] = decharge.date_generation.strftime("%d/%m/%Y")
        ws[CELL_SERVICE] = service_nom

        # ── 4. Fill article table rows ────────────────────────────────────────
        for idx, ligne in enumerate(decharge.lignes.all(), start=1):
            row = ROW_START + idx - 1

            # col A — sequential index
            ws.cell(row=row, column=COL_INDEX, value=idx)

            # col B — designation
            ws.cell(
                row=row,
                column=COL_DESIGNATION,
                value=ligne.id_ressource.designation,
            )

            # col C — quantity
            ws.cell(row=row, column=COL_QUANTITE, value=ligne.quantite)

            # col D — inventory number for bien_inventaire; blank for consommable
            inventaire_num: str = ""
            if (
                ligne.type_ligne == "bien_inventaire"
                and ligne.id_instance_ressource is not None
            ):
                inventaire_num = ligne.id_instance_ressource.numero_inventaire
            ws.cell(row=row, column=COL_INVENTAIRE, value=inventaire_num)

        # ── 5. Save filled workbook to a temp file in /tmp/ ──────────────────
        tmp_dir = Path(tempfile.gettempdir())
        unique_stem = f"decharge_{decharge_id}_{uuid.uuid4().hex}"
        xlsx_tmp_path = tmp_dir / f"{unique_stem}.xlsx"
        wb.save(str(xlsx_tmp_path))
        wb.close()

        # ── 6. Convert .xlsx → PDF via LibreOffice headless ──────────────────
        pdf_dir = tmp_dir  # LibreOffice outputs the PDF next to the xlsx

        result = subprocess.run(  # noqa: S603 — argv fully controlled
            [
                "libreoffice",
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                str(pdf_dir),
                str(xlsx_tmp_path),
            ],
            capture_output=True,
            text=True,
            timeout=60,
            check=False,  # we check returncode manually for a better error msg
        )

        if result.returncode != 0:
            raise RuntimeError(
                f"LibreOffice conversion failed (rc={result.returncode}): "
                f"{result.stderr.strip()}"
            )

        pdf_tmp_path = pdf_dir / f"{unique_stem}.pdf"
        if not pdf_tmp_path.is_file():
            raise FileNotFoundError(
                f"LibreOffice did not produce: {pdf_tmp_path}"
            )

        # ── 7. Move PDF to MEDIA_ROOT/decharges/pdf/ ─────────────────────────
        media_pdf_dir = Path(settings.MEDIA_ROOT) / "decharges" / "pdf"
        media_pdf_dir.mkdir(parents=True, exist_ok=True)

        pdf_filename = f"decharge_{decharge.numero_decharge}.pdf"
        pdf_dest_path = media_pdf_dir / pdf_filename
        shutil.move(str(pdf_tmp_path), str(pdf_dest_path))

        # ── 8. Persist the PDF path on Decharge ──────────────────────────────
        # FileField stores paths relative to MEDIA_ROOT.
        relative_path = os.path.join("decharges", "pdf", pdf_filename)
        Decharge.objects.filter(pk=decharge_id).update(fichier_pdf=relative_path)
        # Refresh the local instance so subsequent code sees the new value
        decharge.refresh_from_db(fields=["fichier_pdf"])

        # ── 9. Create SignatureDecharge (idempotent) ──────────────────────────
        sig, created = SignatureDecharge.objects.get_or_create(
            id_decharge=decharge,
            defaults={
                "statut": "en_attente",
                "id_chef_service": chef,
            },
        )
        if not created and sig.statut not in ("signe", "valide"):
            # Task re-run: reset to en_attente so the chef can (re-)sign
            SignatureDecharge.objects.filter(pk=sig.pk).update(statut="en_attente")

        # ── 10. Notify the chef_service ───────────────────────────────────────
        if chef is not None:
            content_type = ContentType.objects.get_for_model(Decharge)
            Notification.objects.create(
                id_destinataire=chef,
                type_notification="decharge_prete",
                titre="Décharge prête pour signature",
                message=(
                    f"La décharge {decharge.numero_decharge} est disponible "
                    "et attend votre signature."
                ),
                canal="web",
                content_type=content_type,
                object_id=decharge_id,
            )

        logger.info(
            "generate_decharge_pdf: décharge %s → PDF generated (%s)",
            decharge.numero_decharge,
            relative_path,
        )

    except Exception as exc:
        logger.exception(
            "generate_decharge_pdf: error on decharge_id=%s", decharge_id
        )
        countdown = (2 ** self.request.retries) * 60  # 60 s → 120 s
        try:
            raise self.retry(exc=exc, countdown=countdown)
        except MaxRetriesExceededError:
            logger.error(
                "generate_decharge_pdf: max retries exceeded for decharge_id=%s",
                decharge_id,
            )
            # Attempt a best-effort lookup to pass as reference object
            ref = None
            try:
                from apps.decharge.models import Decharge as _D  # noqa: PLC0415
                ref = _D.objects.filter(pk=decharge_id).first()
            except Exception:
                pass

            _notify_gestionnaires(
                msg_titre="Échec génération PDF décharge",
                msg_body=(
                    f"La génération du PDF pour la décharge #{decharge_id} "
                    f"a échoué après plusieurs tentatives. "
                    f"Erreur : {exc}"
                ),
                ref_obj=ref,
            )

    finally:
        # ── 11. Clean up temp xlsx (always) ──────────────────────────────────
        if xlsx_tmp_path is not None and xlsx_tmp_path.is_file():
            try:
                xlsx_tmp_path.unlink()
            except OSError:
                logger.warning(
                    "generate_decharge_pdf: could not delete temp file %s",
                    xlsx_tmp_path,
                )
