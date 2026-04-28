import logging
import os
from decimal import Decimal

from celery.exceptions import MaxRetriesExceededError
from django.db import transaction

from config.celery import app
from apps.procurement.services.ai_extractor import AIExtractor

logger = logging.getLogger(__name__)


def _mark_rejected(import_id: int, observations: str | None = None) -> None:
    from apps.procurement.models import ImportExcelBC  # noqa: PLC0415

    try:
        with transaction.atomic():
            imp = ImportExcelBC.objects.select_for_update().filter(pk=import_id).first()
            if not imp:
                return
            imp.statut_import = "non_conforme"
            if observations:
                imp.observations = observations[:2000]
                imp.save(update_fields=["statut_import", "observations"])
            else:
                imp.save(update_fields=["statut_import"])
            logger.info("[TASK] import %s marked as non_conforme", import_id)
    except Exception:
        logger.exception("[TASK] could not mark import %s as rejected", import_id)


@app.task(queue="ocr", bind=True, max_retries=3)
def extract_excel_items(self, import_id: int, retry_enabled: bool = True) -> None:
    logger.info("[TASK] START extract_excel_items: import_id=%s", import_id)

    from apps.procurement.models import ImportExcelBC, StagingItem  # noqa: PLC0415

    try:
        with transaction.atomic():
            import_obj = ImportExcelBC.objects.select_for_update().get(pk=import_id)
            import_obj.statut_import = "en_attente"
            import_obj.save(update_fields=["statut_import"])

        file_path = import_obj.fichier_excel_original.path
        logger.info("[TASK] Opening Excel file: %s", file_path)

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        import openpyxl  # noqa: PLC0415

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            logger.info("[TASK] Sheet title=%s max_rows=%s", ws.title, ws.max_row)

            lines: list[str] = []
            for row in ws.iter_rows(values_only=True):
                row_values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if not row_values:
                    continue
                line = " | ".join(row_values)
                lines.append(line)

            raw_text = "\n".join(lines)
            logger.info("Raw text size: %s", len(raw_text))

            if not raw_text.strip():
                raise ValueError("Excel extraction produced empty raw text")

            try:
                items = AIExtractor.extract_from_text(raw_text)
            except Exception as exc:
                logger.exception("Library-based extraction failed")
                raise RuntimeError("AI extraction failed") from exc

            logger.info("AI returned %s items", len(items))
            metadata = AIExtractor.build_import_metadata(items)
            if not metadata.get("titre_fichier"):
                metadata["titre_fichier"] = os.path.splitext(os.path.basename(file_path))[0][:255]
            for field_name, field_value in metadata.items():
                setattr(import_obj, field_name, field_value)
            import_obj.save(
                update_fields=[
                    "titre_fichier",
                    "reference_document",
                    "fournisseur_denomination",
                    "fournisseur_telephone",
                    "fournisseur_email",
                    "fournisseur_adresse",
                    "delai_execution",
                ]
            )

            if not items:
                raise ValueError("No items extracted from Excel")

            staging_items: list[StagingItem] = []
            for item in items.get("lignes", []):
                designation = str(item.get("designation", "")).strip()
                description = str(item.get("description", "")).strip()
                if not designation:
                    continue

                quantity = AIExtractor._coerce_int(item.get("quantite", item.get("quantity", 1)), default=1)

                # Apply NLP normalization to get category suggestions
                from apps.procurement.tasks.nlp_normalizer import normalize_designation
                normalized = normalize_designation(designation)

                # Resolve category instance if ID provided
                categorie_instance = None
                if normalized.get("id_categorie_suggeree"):
                    from apps.resources.models import Categorie
                    try:
                        categorie_instance = Categorie.objects.get(id_categorie=normalized["id_categorie_suggeree"])
                    except Categorie.DoesNotExist:
                        categorie_instance = None

                # Resolve resource instance if ID provided
                ressource_instance = None
                if normalized.get("id_ressource_liee"):
                    from apps.resources.models import Ressource
                    try:
                        ressource_instance = Ressource.objects.get(id_ressource=normalized["id_ressource_liee"])
                    except Ressource.DoesNotExist:
                        ressource_instance = None

                staging_items.append(
                    StagingItem(
                        id_import=import_obj,
                        designation_brute=designation[:500],
                        description=description[:4000],
                        designation_normalisee=normalized["designation_normalisee"],
                        quantite=quantity,
                        unite=str(item.get("unite") or "U"),
                        prix_unitaire_ht=item.get("prix_unitaire_ht"),
                        prix_total_ht=item.get("prix_total_ht"),
                        type_detecte=normalized.get("type_detecte", ""),
                        id_categorie_suggeree=categorie_instance,
                        categorie_suggeree_nom=normalized.get("categorie_suggeree_nom", ""),
                        sous_categorie_suggeree_nom=normalized.get("sous_categorie_suggeree_nom", ""),
                        id_ressource_liee=ressource_instance,
                        statut="en_attente",
                    )
                )

            if not staging_items:
                raise ValueError("No valid staging items after validation")

            StagingItem.objects.bulk_create(staging_items, batch_size=500)
            logger.info("[TASK] Created staging items: %s", len(staging_items))

            with transaction.atomic():
                import_final = ImportExcelBC.objects.select_for_update().get(pk=import_id)
                import_final.statut_import = "en_attente"
                import_final.save(update_fields=["statut_import"])

            logger.info("[TASK] DONE extract_excel_items import_id=%s", import_id)
        finally:
            wb.close()

    except Exception as exc:
        logger.exception("[TASK FAILED] import_id=%s error=%s", import_id, exc)

        if not retry_enabled:
            _mark_rejected(import_id, f"Extraction échouée: {str(exc)[:1800]}")
            return

        countdown = (2 ** self.request.retries) * 60
        try:
            raise self.retry(exc=exc, countdown=countdown)
        except MaxRetriesExceededError:
            _mark_rejected(import_id, f"Extraction échouée: {str(exc)[:1800]}")
