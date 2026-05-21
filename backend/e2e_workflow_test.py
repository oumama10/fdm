import os
import sys
from io import BytesIO
from pathlib import Path
from uuid import uuid4

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings.development")
sys.path.insert(0, str(Path(__file__).parent))

import django
django.setup()

from django.core.files.uploadedfile import SimpleUploadedFile
from django.contrib.contenttypes.models import ContentType
from rest_framework.test import APIClient
from openpyxl import Workbook

from apps.users.models import Utilisateur, Role, Service
from apps.resources.models import TypeArticle, Ressource, Stock
from apps.procurement.models import MarcheBC, ImportExcelBC, StagingItem
from apps.procurement.tasks import extract_excel_items
from apps.alerts.models import Notification


# Avoid Redis dependency: execute extraction immediately when view calls .delay()
extract_excel_items.delay = lambda import_id: extract_excel_items(import_id)


def make_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Articles"
    ws["A1"] = "Designation"
    ws["B1"] = "Quantite"
    ws["A2"] = "papier a4"
    ws["B2"] = 12
    ws["A3"] = "stylo bleu"
    ws["B3"] = 5

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()


def setup_data():
    role_fin, _ = Role.objects.get_or_create(nom_role="service_financiere")
    service, _ = Service.objects.get_or_create(
        nom_service="Service E2E",
        defaults={"type_service": "administratif", "description": "E2E tests"},
    )

    user, _ = Utilisateur.objects.get_or_create(
        email="e2e_fin@test.local",
        defaults={
            "nom_complet": "E2E Finance",
            "actif": True,
            "id_role": role_fin,
            "id_service": service,
        },
    )
    user.id_role = role_fin
    user.id_service = service
    user.actif = True
    user.set_password("test123")
    user.save()

    cat = TypeArticle.objects.filter(nom_categorie="consommable").first()
    if cat is None:
        cat = TypeArticle.objects.create(
            nom_categorie="consommable", description="E2E cat"
        )
    res, _ = Ressource.objects.get_or_create(
        designation="papier a4",
        defaults={"id_type": cat, "description": "E2E resource", "unite_mesure": "ramette"},
    )
    Stock.objects.get_or_create(
        id_ressource=res,
        defaults={"quantite_disponible": 100, "seuil_alerte": 5},
    )

    marche = MarcheBC.objects.create(
        reference=f"REF-E2E-{uuid4().hex[:8].upper()}",
        type_acquisition="marche",
        statut="en_attente_livraison",
        id_cree_par=user,
    )
    return user, marche


def main():
    print("=== E2E HTTP Upload + Extraction Test ===")
    user, marche = setup_data()

    file_content = make_workbook_bytes()
    upload_file = SimpleUploadedFile(
        "e2e_import.xlsx",
        file_content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    client = APIClient()
    client.force_authenticate(user=user)

    response = client.post(
        "/api/procurement/import/",
        {
            "fichier_excel_original": upload_file,
            "source_type": "bc",
            "id_marche": marche.id_marche,
            "observations": "e2e",
        },
        format="multipart",
    )

    print(f"POST /api/procurement/import/ -> {response.status_code}")
    if response.status_code not in (200, 201):
        print(response.content.decode(errors="ignore"))
        raise SystemExit(1)

    payload = response.json()
    import_id = payload.get("id_import") or payload.get("id") or payload.get("idImport")
    if import_id is None:
        print("Unexpected response payload:", payload)
        raise SystemExit(1)
    print(f"Import ID: {import_id}")

    detail = client.get(f"/api/procurement/import/{import_id}/")
    print(f"GET /api/procurement/import/{import_id}/ -> {detail.status_code}")

    items = StagingItem.objects.filter(id_import_id=import_id).order_by("id_staging")
    notifs = Notification.objects.filter(
        content_type=ContentType.objects.get_for_model(ImportExcelBC),
        object_id=import_id,
    )

    print(f"Staging items: {items.count()}")
    for row in items:
        linked = row.id_ressource_liee.designation if row.id_ressource_liee else "None"
        print(
            f"- {row.designation_brute} | norm={row.designation_normalisee} | qte={row.quantite} | conf={row.confiance_ia} | linked={linked}"
        )

    print(f"Notifications: {notifs.count()}")
    print("RESULT:", "PASS" if items.count() >= 2 else "FAIL")


if __name__ == "__main__":
    main()
